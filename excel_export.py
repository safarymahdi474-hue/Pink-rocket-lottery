import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import database as db


async def export_users_excel() -> io.BytesIO:
    participants = await db.get_all_participants()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "شرکت‌کنندگان"
    ws.sheet_view.rightToLeft = True

    headers = ["User ID", "Username", "نام", "تعداد تیکت", "تعداد رفرال"]
    header_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (user_id, tickets, referrals, username, full_name) in enumerate(participants, 2):
        ws.cell(row=row_idx, column=1, value=user_id)
        ws.cell(row=row_idx, column=2, value=f"@{username}" if username else "-")
        ws.cell(row=row_idx, column=3, value=full_name or "-")
        ws.cell(row=row_idx, column=4, value=tickets)
        ws.cell(row=row_idx, column=5, value=referrals)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
